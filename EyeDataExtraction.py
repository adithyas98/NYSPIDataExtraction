#!/usr/bin/env python


# By Adithya Shastry
# This script will extract the eye tracking data from an 
				#excel file and will upload
# to Redcap using the RedCap Api python class

#import stuff
from redcapAPI import RedCapAPI as red
from openpyxl import load_workbook #this is the main library 
						#needed to get data from the Excel File
import os
#Store the directory that we want to look into
directory = "./data/"



def extractData(sheet):
	#This method will extract the data
	#it requires a sheet item to be sent
	
	#first extract the labels
	rawLabels = sheet[1]
	#We need to extract the values from the output
	labels = [] 
	for l in rawLabels:
		labels.append(l.value)	
	# Now we can iterate through each of the rows and store them
	for row in sheet.values:
		print(row)
		#TODO: We can process this now with the redcap API class

for filename in os.listdir(directory):
	if (filename.endswith(".xlsx")):
		#we want to call our data extraction method
		#first we want to load the workbook
		workbook = load_workbook(filename=directory+filename)
		#now we want to iterate through each of the sheets
		sheets = workbook.sheetnames
		for sheet in sheets:
			extractData(workbook[sheet])

